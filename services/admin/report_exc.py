from flask import jsonify, request, render_template, send_file
from app import mongo
from datetime import datetime, timedelta
from calendar import month_name
import pytz
from openpyxl.utils import get_column_letter
from bson import ObjectId
from io import BytesIO
import openpyxl
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def get_revenue_report(time_frame='month'):
    """
    Tạo báo cáo doanh thu và lợi nhuận dựa trên khung thời gian: month, quarter, year.
    """
    try:
        # Xác định năm hiện tại
        selected_year = datetime.now().year

        # Xác định khoảng thời gian
        start_date = datetime(selected_year, 1, 1)
        end_date = datetime(selected_year, 12, 31, 23, 59, 59)

        # Tính tổng số đơn hàng
        total_orders = mongo.db.orders.count_documents({})

        # Tính trạng thái đơn hàng
        statuses = ["pending", "waiting_for_shipping", "waiting_for_delivery", "completed", "cancelled", "returned"]
        status_counts = {status: mongo.db.orders.count_documents({"delivery_status": status}) for status in statuses}
        total_status = sum(status_counts.values())
        status_data = [status_counts[status] for status in statuses]
        status_percentages = [(count / total_status * 100) if total_status > 0 else 0 for count in status_data]

        # Xác định số khoảng thời gian dựa trên time_frame
        if time_frame == 'month':
            num_intervals = 12
        elif time_frame == 'quarter':
            num_intervals = 4
        else:  # year
            num_intervals = 1

        # Chuyển start_date và end_date thành chuỗi để so sánh với order_date
        start_date_str = start_date.strftime('%Y-%m-%d %H:%M:%S')
        end_date_str = end_date.strftime('%Y-%m-%d %H:%M:%S')

        # Pipeline để tính doanh thu và lợi nhuận
        pipeline = [
            # Lọc đơn hàng trong khoảng thời gian và trạng thái "completed"
            {
                "$match": {
                    "order_date": {
                        "$gte": start_date_str,
                        "$lte": end_date_str
                    },
                    "delivery_status": "completed"  # Chỉ lấy đơn hàng đã hoàn thành
                }
            },
            # Unwind items để xử lý từng sản phẩm trong đơn hàng
            {"$unwind": "$items"},
            # Truy vấn thông tin sản phẩm từ collection products
            {
                "$lookup": {
                    "from": "products",
                    "let": {"item_id": "$items._id"},
                    "pipeline": [
                        {
                            "$match": {
                                "$expr": {
                                    "$eq": ["$_id", {"$toObjectId": "$$item_id"}]
                                }
                            }
                        }
                    ],
                    "as": "product_info"
                }
            },
            # Unwind product_info để xử lý dữ liệu sản phẩm
            {"$unwind": "$product_info"},
            # Tính doanh thu và lợi nhuận cho từng sản phẩm
            {
                "$project": {
                    "order_date": 1,
                    "quantity": "$items.quantity",
                    "discounted_price": "$product_info.discounted_price",
                    "cost_price": "$product_info.cost_price",
                    "revenue": {
                        "$multiply": [
                            {"$toDouble": "$product_info.discounted_price"},
                            {"$toInt": "$items.quantity"}
                        ]
                    },
                    "profit": {
                        "$multiply": [
                            {
                                "$subtract": [
                                    {"$toDouble": "$product_info.discounted_price"},
                                    {"$toDouble": "$product_info.cost_price"}
                                ]
                            },
                            {"$toInt": "$items.quantity"}
                        ]
                    }
                }
            }
        ]

        # Thêm bước nhóm theo time_frame
        if time_frame == "month":
            pipeline.append({
                "$group": {
                    "_id": {
                        "year": {"$year": {"$dateFromString": {"dateString": "$order_date"}}},
                        "month": {"$month": {"$dateFromString": {"dateString": "$order_date"}}}
                    },
                    "total_revenue": {"$sum": "$revenue"},
                    "total_profit": {"$sum": "$profit"}
                }
            })
            pipeline.append({"$match": {"_id.year": selected_year}})
            pipeline.append({"$sort": {"_id.month": 1}})

        elif time_frame == "quarter":
            pipeline.append({
                "$group": {
                    "_id": {
                        "year": {"$year": {"$dateFromString": {"dateString": "$order_date"}}},
                        "quarter": {
                            "$ceil": {
                                "$divide": [
                                    {"$month": {"$dateFromString": {"dateString": "$order_date"}}},
                                    3
                                ]
                            }
                        }
                    },
                    "total_revenue": {"$sum": "$revenue"},
                    "total_profit": {"$sum": "$profit"}
                }
            })
            pipeline.append({"$match": {"_id.year": selected_year}})
            pipeline.append({"$sort": {"_id.quarter": 1}})

        else:  # year
            pipeline.append({
                "$group": {
                    "_id": {
                        "year": {"$year": {"$dateFromString": {"dateString": "$order_date"}}}
                    },
                    "total_revenue": {"$sum": "$revenue"},
                    "total_profit": {"$sum": "$profit"}
                }
            })
            pipeline.append({"$match": {"_id.year": selected_year}})
            pipeline.append({"$sort": {"_id.year": 1}})

        # Thực thi pipeline
        try:
            data = list(mongo.db.orders.aggregate(pipeline))
            print(f"Data: {data}")  # Debug
        except Exception as e:
            print(f"Error in aggregation pipeline: {str(e)}")
            data = []

        # Chuẩn bị dữ liệu cho biểu đồ
        labels = []
        revenue_values = []
        profit_values = []
        total_revenue = 0

        if time_frame == "month":
            for month in range(1, 13):
                labels.append(month_name[month])
                revenue_values.append(0)
                profit_values.append(0)
            for entry in data:
                month = entry["_id"]["month"]
                revenue_values[month - 1] = entry["total_revenue"]
                profit_values[month - 1] = entry["total_profit"]
                total_revenue += entry["total_revenue"]

        elif time_frame == "quarter":
            for quarter in range(1, 5):
                labels.append(f"Q{quarter}")
                revenue_values.append(0)
                profit_values.append(0)
            for entry in data:
                quarter = entry["_id"]["quarter"]
                revenue_values[quarter - 1] = entry["total_revenue"]
                profit_values[quarter - 1] = entry["total_profit"]
                total_revenue += entry["total_revenue"]

        else:  # year
            labels.append(str(selected_year))
            revenue_values.append(0)
            profit_values.append(0)
            for entry in data:
                revenue_values[0] = entry["total_revenue"]
                profit_values[0] = entry["total_profit"]
                total_revenue = entry["total_revenue"]

        # Tạo dictionary chứa dữ liệu báo cáo
        report_data = {
            'time_frame': time_frame,
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'status_chart': {
                'labels': ['Đang chuẩn bị', 'Chờ vận chuyển', 'Chờ giao hàng', 'Hoàn thành', 'Đã hủy', 'Hoàn hàng'],
                'data': status_data,
                'percentages': status_percentages
            },
            'revenue_chart': {
                'labels': labels,
                'data': revenue_values
            },
            'profit_chart': {
                'labels': labels,
                'data': profit_values
            },
            'start_date': start_date_str,
            'end_date': end_date_str
        }

        print(f"Report data: {report_data}")  # Debug
        return report_data

    except Exception as e:
        print(f"Error in get_revenue_report: {str(e)}")
        # Trả về dữ liệu mặc định nếu có lỗi
        return {
            'time_frame': time_frame,
            'total_orders': 0,
            'total_revenue': 0,
            'status_chart': {
                'labels': ['Đang chuẩn bị', 'Chờ vận chuyển', 'Chờ giao hàng', 'Hoàn thành', 'Đã hủy', 'Hoàn hàng'],
                'data': [0, 0, 0, 0, 0, 0],
                'percentages': [0, 0, 0, 0, 0, 0]
            },
            'revenue_chart': {
                'labels': [],
                'data': []
            },
            'profit_chart': {
                'labels': [],
                'data': []
            },
            'start_date': '',
            'end_date': ''
        }

def export_revenue_report_to_excel():
    """
    Xuất báo cáo doanh thu và lợi nhuận ra file Excel, bao gồm biểu đồ.
    """
    try:
        # Lấy time_frame từ query string (nếu có), mặc định là 'month'
        time_frame = request.args.get('time_frame', 'month')
        
        # Lấy dữ liệu báo cáo bằng cách gọi hàm get_revenue_report
        report_data = get_revenue_report(time_frame)

        # Tạo workbook và worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Revenue Report"

        # Định dạng chung
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Thêm tiêu đề chính
        worksheet['A1'] = "BÁO CÁO DOANH THU"
        worksheet['A1'].font = title_font
        worksheet['A1'].alignment = center_alignment
        worksheet.merge_cells('A1:I1')

        # Thêm thông tin tổng quan
        worksheet['A3'] = "Tổng số đơn hàng"
        worksheet['B3'] = report_data['total_orders']
        worksheet['A4'] = "Tổng doanh thu (VNĐ)"
        worksheet['B4'] = report_data['total_revenue']
        worksheet['A5'] = "Khung thời gian"
        worksheet['B5'] = report_data['time_frame'].capitalize()
        worksheet['A6'] = "Từ ngày"
        worksheet['B6'] = report_data['start_date']
        worksheet['A7'] = "Đến ngày"
        worksheet['B7'] = report_data['end_date']

        # Định dạng cột thông tin tổng quan
        for row in range(3, 8):
            worksheet[f'A{row}'].font = header_font
            worksheet[f'A{row}'].border = border
            worksheet[f'B{row}'].border = border
            worksheet[f'A{row}'].alignment = center_alignment
            worksheet[f'B{row}'].alignment = center_alignment

        # Thêm dữ liệu trạng thái đơn hàng
        worksheet['A9'] = "Trạng thái đơn hàng"
        worksheet['A9'].font = header_font
        worksheet['A9'].alignment = center_alignment
        worksheet.merge_cells('A9:C9')
        worksheet['A10'] = "Trạng thái"
        worksheet['B10'] = "Số lượng"
        worksheet['C10'] = "Tỷ lệ (%)"
        for idx, (label, count, percentage) in enumerate(zip(report_data['status_chart']['labels'], report_data['status_chart']['data'], report_data['status_chart']['percentages']), start=11):
            worksheet[f'A{idx}'] = label
            worksheet[f'B{idx}'] = count
            worksheet[f'C{idx}'] = round(percentage, 2)

        # Định dạng bảng trạng thái đơn hàng
        for row in range(9, 17):
            for col in ['A', 'B', 'C']:
                cell = worksheet[f'{col}{row}']
                cell.border = border
                cell.alignment = center_alignment
        for col in ['A', 'B', 'C']:
            worksheet[f'{col}10'].font = header_font

        # Thêm biểu đồ trạng thái đơn hàng (Pie Chart)
        pie_chart = PieChart()
        labels = Reference(worksheet, min_col=1, min_row=11, max_row=16)  # Cột "Trạng thái"
        data = Reference(worksheet, min_col=2, min_row=11, max_row=16)  # Cột "Số lượng"
        pie_chart.add_data(data, titles_from_data=False)
        pie_chart.set_categories(labels)
        pie_chart.title = "Trạng thái đơn hàng"
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showCatName = True  # Hiển thị tên trạng thái
        pie_chart.dataLabels.showPercent = True  # Hiển thị phần trăm
        pie_chart.dataLabels.showVal = False  # Không hiển thị giá trị số lượng
        pie_chart.dataLabels.showSerName = False  # Không hiển thị "Series1", "Series2",...
        # Đặt màu sắc cho các phần của biểu đồ
        colors = [
            '808080',  # Đang chuẩn bị - Gray
            'DC3545',  # Chờ vận chuyển - Red
            'FFC107',  # Chờ giao hàng - Yellow
            '2CE64E',  # Hoàn thành - Green
            'FF0000',  # Đã hủy - Red (darker shade)
            'E0C846'   # Hoàn hàng - Yellow (brighter shade)
        ]
        for idx, series in enumerate(pie_chart.series):
            for i, point in enumerate(series.dPt):
                point.graphicalProperties.solidFill = colors[i]
        worksheet.add_chart(pie_chart, "E9")  # Đặt biểu đồ ở vị trí E9

        # Thêm dữ liệu doanh thu và lợi nhuận
        row_start = 18
        worksheet[f'A{row_start}'] = "Doanh thu và lợi nhuận theo thời gian"
        worksheet[f'A{row_start}'].font = header_font
        worksheet[f'A{row_start}'].alignment = center_alignment
        worksheet.merge_cells(f'A{row_start}:C{row_start}')
        worksheet['A19'] = "Thời gian"
        worksheet['B19'] = "Doanh thu (VNĐ)"
        worksheet['C19'] = "Lợi nhuận (VNĐ)"
        for idx, (label, revenue, profit) in enumerate(zip(report_data['revenue_chart']['labels'], report_data['revenue_chart']['data'], report_data['profit_chart']['data']), start=20):
            worksheet[f'A{idx}'] = label
            worksheet[f'B{idx}'] = revenue
            worksheet[f'C{idx}'] = profit

        # Định dạng bảng doanh thu và lợi nhuận
        for row in range(18, 32):
            for col in ['A', 'B', 'C']:
                cell = worksheet[f'{col}{row}']
                cell.border = border
                cell.alignment = center_alignment
        for col in ['A', 'B', 'C']:
            worksheet[f'{col}19'].font = header_font

        # Thêm biểu đồ doanh thu theo thời gian (Line Chart)
        if time_frame == "month":  # Chỉ tạo biểu đồ đường cho khung thời gian "month"
            line_chart = LineChart()
            line_chart.title = "Doanh thu theo tháng"
            line_chart.y_axis.title = "Doanh thu (VNĐ)"
            line_chart.x_axis.title = "Tháng"

            # Dữ liệu doanh thu (bao gồm tiêu đề "Doanh thu (VNĐ)" để dùng làm tên series)
            data = Reference(worksheet, min_col=2, min_row=19, max_row=31)  # Cột "Doanh thu (VNĐ)", bao gồm header
            line_chart.add_data(data, titles_from_data=True)  # titles_from_data=True để lấy tiêu đề từ ô B19

            # Nhãn trục x (Tháng)
            labels = Reference(worksheet, min_col=1, min_row=20, max_row=31)  # Cột "Thời gian"
            line_chart.set_categories(labels)

            # Định dạng trục y để hiển thị tiền tệ
            line_chart.y_axis.number_format = '#,##0'
            line_chart.y_axis.scaling.min = 0  # Bắt đầu trục y từ 0
            line_chart.y_axis.majorGridlines = None  # Tắt gridlines chính nếu không cần
            line_chart.x_axis.tickLblPos = "low"  # Đặt nhãn trục x ở dưới

            # Đặt màu và độ dày cho đường biểu đồ
            line_chart.series[0].graphicalProperties.line.solidFill = '007BFF'  # Màu xanh
            line_chart.series[0].graphicalProperties.line.width = 25000  # Độ dày đường (2.5 pt)

            # Thêm nhãn dữ liệu (data labels) trên các điểm
            line_chart.dataLabels = DataLabelList()
            line_chart.dataLabels.showVal = True  # Hiển thị giá trị doanh thu
            line_chart.dataLabels.number_format = '#,##0'  # Định dạng tiền tệ cho nhãn

            worksheet.add_chart(line_chart, "E20")  # Đặt biểu đồ ở vị trí E20

        # Điều chỉnh độ rộng cột cho đẹp
        column_widths = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:  # Chỉ xử lý các ô có giá trị
                    column_letter = get_column_letter(cell.column)
                    column_widths[column_letter] = max(
                        column_widths.get(column_letter, 0),
                        len(str(cell.value))
                    )

        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = width + 2

        # Lưu file vào BytesIO để trả về
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Trả về file Excel để tải xuống
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'revenue_report_{time_frame}.xlsx'
        )

    except Exception as e:
        print(f"Error in export_revenue_report_to_excel: {str(e)}")
        return jsonify({"error": "Không thể xuất file Excel", "message": str(e)}), 500